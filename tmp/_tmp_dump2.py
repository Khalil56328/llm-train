# -*- coding: utf-8 -*-
"""临时脚本：完整读取功能点 xlsx 每个单元格"""
import os
from openpyxl import load_workbook
c = r'd:\work\project\20260806\docs\训推平台功能点.xlsx'
wb = load_workbook(c, data_only=True)
for ws in wb.worksheets:
    print('SHEET:', ws.title, 'dims:', ws.dimensions)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                print('--- %s%s:' % (cell.coordinate, ' (merged)' if cell.coordinate in [str(m) for m in ws.merged_cells.ranges] else ''))
                print(repr(str(cell.value)))
